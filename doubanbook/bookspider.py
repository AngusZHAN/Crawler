import os
import requests
from openpyxl import Workbook
from pyquery import PyQuery as pq

class Model(object):
    '''
    基类, 用来显示类的信息
    '''
    def __repr__(self):
        name = self.__class__.__name__
        properties = ('{}=({})'.format(k, v) for k, v in self.__dict__.items())
        s = '\n<{} \n  {}>'.format(name, '\n  '.join(properties))
        return s


class Book(Model):
    '''
    存储书籍信息
    '''
    def __init__(self):
        self.name = ''
        self.score = 0
        self.cover_url = ''
        self.ranking = 0


def cached_url(url, t='无'):
    '''
    网页缓存, 避免重复下载网页
    '''
    folder = 'cached'
    filename = url.split('=', 1)[-1] + t + '.html'
    path = os.path.join(folder, filename)
    if os.path.exists(path):
        with open(path, 'rb') as f:
            s = f.read()
            return s
    else:
        # 建立 cached 文件夹
        if not os.path.exists(folder):
            os.makedirs(folder)

        headers = {
            'user-agent': '''Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.98 Safari/537.36
             Accept: text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8''',
        }
        # 发送网络请求, 把结果写入到文件夹中
        r = requests.get(url, headers)
        with open(path, 'wb') as f:
            f.write(r.content)
        return r.content


def book_from_div(div):
    '''
    从一个 div 里面获取到一个书籍信息
    '''
    e = pq(div)

    # 小作用域变量用单字符
    b = Book()
    b.name = e('h2').text()
    b.score = e('.rating_nums').text()
    b.pub = e('.pub').text
    b.cover_url = e('img').attr('src')

    return b


def book_from_url(url, t):
    '''
    从 url 中下载网页并解析出页面内所有的书籍
    只会下载一次
    '''
    page = cached_url(url, t)
    '''
    1. 解析 dom
    2. 找到父亲节点
    3. 每个子节点拿一个movie
    '''
    e = pq(page)
    # print(page.decode())
    # 2.父节点
    items = e('.subject-item')
    # 调用 book_from_div
    # list comprehension
    books = [book_from_div(i) for i in items]
    return books


def download_image(url):
    folder = "img"
    name = url.split("/")[-1]
    path = os.path.join(folder, name)

    if not os.path.exists(folder):
        os.makedirs(folder)

    if os.path.exists(path):
        return

    headers = {
        'user-agent': '''Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.98 Safari/537.36
         Accept: text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8''',
    }
    # 发送网络请求, 把结果写入到文件夹中
    r = requests.get(url, headers)
    with open(path, 'wb') as f:
        f.write(r.content)


def write_excel(tag, books):
    '''
    将爬去的书籍信息写入excel文件
    '''
    wb = Workbook(optimized_write=True)
    ws = []
    ws.append(wb.create_sheet(title=tag.decode())) #utf8->unicode
    
    for i in range(len(tags)):
        ws[i].append(['排名', '书名', '评分', '出版信息'])
        count = 1
        for b in books:
            ws[i].append([count, b[0], float(b[1]), b[2]])
            count += 1
    
    save_path = 'book_list'
    save_path += ('-' + tag.decode())
    save_path += '.xlsx'
    wb.save(save_path)


def main():
    '''
    爬取每个标签下评价前100本书
    '''
    tags = ['哲学', '历史', '金融', '心理学', '传记']
    for tag in tags:
        for i in range(0, 100, 20):
            url = 'https://book.douban.com/tag/{}?start={}&type=S'.format(t, i)
            books = book_from_url(url, tag)
            print(tag)
            print('书单', books)
            write_excel(tag, books)
            [download_image(b.cover_url) for b in books]


if __name__ == '__main__':
    main()
